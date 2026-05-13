import os
import sys
import django
import openpyxl
from decimal import Decimal
 
# Setup Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nifty100_project.settings')
django.setup()
 
from nifty_api.models import DimCompany, DimYear, FactProfitLoss, FactBalanceSheet
 
def import_profit_loss():
    """Import Profit & Loss data"""
    print("=" * 60)
    print("📊 Importing Profit & Loss Data")
    print("=" * 60)
    
    wb = openpyxl.load_workbook('profitandloss.xlsx')
    sheet = wb.active
    
    count = 0
    skipped = 0
    
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i <= 2:  # Skip title and header
            continue
        
        company_id = row[1]
        year_label = row[2]
        sales = row[3]
        expenses = row[4]
        operating_profit = row[5]
        opm_pct = row[6]
        other_income = row[7]
        interest = row[8]
        depreciation = row[9]
        profit_before_tax = row[10]
        tax_pct = row[11]
        net_profit = row[12]
        eps = row[13]
        dividend_payout = row[14]
        
        if not company_id or not year_label:
            continue
        
        # Get or create company
        try:
            company = DimCompany.objects.get(symbol=company_id)
        except DimCompany.DoesNotExist:
            skipped += 1
            continue
        
        # Get or create year
        year, _ = DimYear.objects.get_or_create(
            year_label=year_label,
            defaults={'fiscal_year': None, 'sort_order': None}
        )
        
        # Create or update P&L record
        FactProfitLoss.objects.update_or_create(
            symbol=company,
            year=year,
            defaults={
                'sales': Decimal(str(sales)) if sales else None,
                'net_profit': Decimal(str(net_profit)) if net_profit else None,
                'opm_pct': Decimal(str(opm_pct)) if opm_pct else None,
                'eps': Decimal(str(eps)) if eps else None,
                'net_profit_margin_pct': None,
            }
        )
        count += 1
        
        if count % 100 == 0:
            print(f"   Processed {count} records...")
    
    print(f"✅ Imported {count} P&L records")
    print(f"⚠️  Skipped {skipped} records (company not found)")
 
def import_balance_sheet():
    """Import Balance Sheet data"""
    print("\n" + "=" * 60)
    print("📊 Importing Balance Sheet Data")
    print("=" * 60)
    
    wb = openpyxl.load_workbook('balancesheet.xlsx')
    sheet = wb.active
    
    count = 0
    skipped = 0
    
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i <= 2:  # Skip title and header
            continue
        
        company_id = row[1]
        year_label = row[2]
        total_assets = row[12]
        borrowings = row[5]
        
        if not company_id or not year_label:
            continue
        
        # Get or create company
        try:
            company = DimCompany.objects.get(symbol=company_id)
        except DimCompany.DoesNotExist:
            skipped += 1
            continue
        
        # Get or create year
        year, _ = DimYear.objects.get_or_create(
            year_label=year_label,
            defaults={'fiscal_year': None, 'sort_order': None}
        )
        
        # Create or update Balance Sheet record
        FactBalanceSheet.objects.update_or_create(
            symbol=company,
            year=year,
            defaults={
                'total_assets': Decimal(str(total_assets)) if total_assets else None,
                'borrowings': Decimal(str(borrowings)) if borrowings else None,
                'debt_to_equity': None,
            }
        )
        count += 1
        
        if count % 100 == 0:
            print(f"   Processed {count} records...")
    
    print(f"✅ Imported {count} Balance Sheet records")
    print(f"⚠️  Skipped {skipped} records (company not found)")
 
def show_summary():
    """Show import summary"""
    print("\n" + "=" * 60)
    print("📈 Import Summary")
    print("=" * 60)
    print(f"Companies: {DimCompany.objects.count()}")
    print(f"Years: {DimYear.objects.count()}")
    print(f"P&L Records: {FactProfitLoss.objects.count()}")
    print(f"Balance Sheet Records: {FactBalanceSheet.objects.count()}")
    
    # Calculate totals
    from django.db.models import Sum, Avg
    
    total_sales = FactProfitLoss.objects.aggregate(Sum('sales'))['sales__sum']
    total_profit = FactProfitLoss.objects.aggregate(Sum('net_profit'))['net_profit__sum']
    avg_opm = FactProfitLoss.objects.aggregate(Avg('opm_pct'))['opm_pct__avg']
    
    print(f"\n💰 Financial Metrics:")
    print(f"Total Sales: ₹{total_sales:,.0f} Cr" if total_sales else "Total Sales: N/A")
    print(f"Total Net Profit: ₹{total_profit:,.0f} Cr" if total_profit else "Total Net Profit: N/A")
    print(f"Average OPM: {avg_opm:.2f}%" if avg_opm else "Average OPM: N/A")
    print("=" * 60)
 
if __name__ == '__main__':
    import_profit_loss()
    import_balance_sheet()
    show_summary()