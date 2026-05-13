import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'nifty100_project.settings')
django.setup()

from nifty_api.models import DimCompany, DimYear, FactProfitLoss, FactBalanceSheet
import openpyxl
from decimal import Decimal

print(f"Current companies in DB: {DimCompany.objects.count()}")

# Load companies if needed
if DimCompany.objects.count() < 92:
    print("Loading companies from Excel...")
    wb = openpyxl.load_workbook('companies.xlsx')
    sheet = wb.active
    
    count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        
        symbol = row[0]
        company_name = row[2]
        website = row[5] if row[5] else ''
        face_value = row[8] if row[8] else 0
        book_value = row[9] if row[9] else 0
        
        if symbol and company_name:
            DimCompany.objects.get_or_create(
                symbol=symbol,
                defaults={
                    'company_name': company_name,
                    'website': website,
                    'face_value': Decimal(str(face_value)) if face_value else None,
                    'book_value': Decimal(str(book_value)) if book_value else None,
                }
            )
            count += 1
    
    print(f"✅ Loaded {count} companies! Total now: {DimCompany.objects.count()}")
else:
    print(f"✅ Database has {DimCompany.objects.count()} companies")

# Load financial data if needed
if FactProfitLoss.objects.count() < 100:
    print("Loading P&L data...")
    wb_pl = openpyxl.load_workbook('profitandloss.xlsx')
    sheet_pl = wb_pl.active
    
    count = 0
    for i, row in enumerate(sheet_pl.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        
        company_id = row[1]
        year_label = row[2]
        sales = row[3]
        net_profit = row[12]
        opm_pct = row[6]
        eps = row[13]
        
        if not company_id or not year_label:
            continue
        
        try:
            company = DimCompany.objects.get(symbol=company_id)
            year, _ = DimYear.objects.get_or_create(year_label=year_label)
            
            FactProfitLoss.objects.get_or_create(
                symbol=company,
                year=year,
                defaults={
                    'sales': Decimal(str(sales)) if sales else None,
                    'net_profit': Decimal(str(net_profit)) if net_profit else None,
                    'opm_pct': Decimal(str(opm_pct)) if opm_pct else None,
                    'eps': Decimal(str(eps)) if eps else None,
                }
            )
            count += 1
        except:
            pass
    
    print(f"✅ Loaded {count} P&L records!")
else:
    print(f"✅ Database has {FactProfitLoss.objects.count()} P&L records")